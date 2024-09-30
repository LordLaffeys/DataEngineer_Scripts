#V1.1 
# 
# (ENHANCED --> auto get json from server dwh & cc)
# New requirements libraries: sqlalchemy, pandas

import requests
import json
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from tqdm import tqdm
from datetime import datetime,time
import urllib3
import pytz
from dotenv import load_dotenv
import os
from collections import defaultdict
import pandas as pd 
from sqlalchemy import create_engine
import sys



def get_token(username, password):
    login_path = "https://edm-delman.apps.binus.edu/analytic/login"

    payload = json.dumps({
        "username": username,
        "password": password
    })
    headers = {
    'Content-Type': 'application/json'
    }

    token = requests.post(
        login_path, 
        headers=headers, 
        data=payload, 
        verify=False
    ).headers["Authorization"]
    
    return token

def convert_timezone_to_gmt7(datetime_str):
    # Define timezones
    utc_timezone = pytz.timezone('Etc/GMT')
    gmt7_timezone = pytz.timezone('Asia/Bangkok')  # GMT+0700 is Indochina Time (ICT)
    
    # Parse the datetime string
    dt = datetime.strptime(datetime_str, '%a, %d %b %Y, %H:%M:%S GMT+0000')
    
    # Localize to UTC timezone
    utc_dt = utc_timezone.localize(dt)
    
    # Convert to GMT+0700 (ICT)
    gmt7_dt = utc_dt.astimezone(gmt7_timezone)
    
    return gmt7_dt.strftime('%a, %d %b %Y, %H:%M:%S GMT+0700')


def get_projects_json(proj_file):
    dwh_username = os.environ.get('DWH_USERNAME')
    dwh_password = os.environ.get('DWH_PASSWORD')
    cc_username = os.environ.get('CC_USERNAME')
    cc_password = os.environ.get('CC_PASSWORD')


    if proj_file.lower() == 'dwh':
        dwh_url = f'mssql+pyodbc://{dwh_username}:{dwh_password}@edm-dwh.binus.db:1433/dwh?driver=ODBC+Driver+17+for+SQL+Server'
        dwh_engine = create_engine(dwh_url)

        # Get json data from table in server
        query = "SELECT [name], [id], [name] as init_name, 'Delman' AS [loc] FROM DailyJobCheck"
        
        # Execute the query and load the result into a DataFrame
        df = pd.read_sql(query, dwh_engine)

        # Convert DataFrame to JSON string
        json_str = df.to_json(orient='records', indent=4)
        return json_str
    
    elif proj_file.lower() == 'cc':
        cc_url = f'mssql+pyodbc://{cc_username}:{cc_password}@edm-comcen.binus.db:1433/CommandCenter_DB?driver=ODBC+Driver+17+for+SQL+Server'
        cc_engine = create_engine(cc_url)

        # Get json data from table in server
        query = "SELECT [name], id, [name] AS init_name, 'Delman' AS [loc] FROM DimDelmanProjectsCC"
        
        # Execute the query and load the result into a DataFrame
        df = pd.read_sql(query, cc_engine)
        
        # Convert DataFrame to JSON string
        json_str = df.to_json(orient='records', indent=4)
        return json_str


def generate_excel(token, excel, proj_file="projects.json"):

    if proj_file.lower() in ['dwh', 'cc']:
        proj_json_str = get_projects_json(proj_file)
        proj_dict = json.loads(proj_json_str)

    else:
        if not os.path.exists(proj_file):
            print(f"The file '{proj_file}' does not exist. try with correct 'file.json' name or use 'dwh' or 'cc'")
            return
         
        with open(proj_file) as f:
            proj_dict = json.load(f)
    
    proj_url = "https://edm-delman.apps.binus.edu/analytic/projects/"

    explored = defaultdict(lambda:None)
    stat = ["SUCCESS", None, "CREATED", "UPSTREAM FAILED"]

    for i, proj in tqdm(enumerate(proj_dict), total=len(proj_dict)):
        error_node, error_note = None, None
        status = ""
        if proj["id"] != None:
            if explored[proj["id"]] is not None:
                note, status = explored[proj["id"]]
                _, _ = excel.write_line(proj, i+2, error_note=note, status=status)
                continue


            schedules = requests.get(
                proj_url+proj["id"]+"/schedules?page_size=8&page=0",
                headers= {'Authorization': token},
                verify=False
            )
            
            # Check the schedules time
            data = json.loads(schedules.content)['data']       
            if data:
                for entry in data:
                    repeat_period = entry.get('repeat_period', {})
                    if entry['repeat_period'] == "beginning_of_the_month" :
                        continue
                    elif 'day_of_week' in repeat_period or 'day' in repeat_period:
                        repeat_period = entry['repeat_period']['hour']
                        time_object = time(hour=repeat_period, minute=0)
                        # Convert to GMT+7
                        gmt_offset = 25  # GMT+7 offset in hours
                        new_hour = (time_object.hour + gmt_offset) % 24  # Calculate new hour accounting for overflow
                        converted_time = time(hour=new_hour, minute=time_object.minute)
                        # Define the comparison time (18:00:00)
                        comparison_time = time(hour=18, minute=0)
                        break
                    else :
                        repeat_period = entry['repeat_period']['hour']
                        time_object = time(hour=repeat_period, minute=0)
                        # Convert to GMT+7
                        gmt_offset = 7  # GMT+7 offset in hours
                        new_hour = (time_object.hour + gmt_offset) % 24  # Calculate new hour accounting for overflow
                        converted_time = time(hour=new_hour, minute=time_object.minute)
                        # Define the comparison time (18:00:00)
                        comparison_time = time(hour=18, minute=0)
                        break


                if converted_time < comparison_time :
                    monitoring = requests.get(
                    proj_url+proj["id"]+"/monitoring?page_size=8&page=0",
                    headers= {'Authorization': token},
                    verify=False
                    )

                        # Check the sync date
                    data = json.loads(monitoring.content)['data']       
                    if data:
                        for entry in data:
                            if entry.get('started_at'):  # Check if 'started_at' exists and is not None
                                date_sync = entry['started_at']
                                gmt7_datetime = convert_timezone_to_gmt7(date_sync)
                                datetime_obj = datetime.strptime(gmt7_datetime, "%a, %d %b %Y, %H:%M:%S %Z%z")
                                date_only = datetime_obj.date()
                                current_date = datetime.now().date()
                                # Perform further operations with date_sync if needed
                                break  # Exit the loop once a valid 'started_at' is found
                
                        if  date_only >= current_date: # check node sync date with the current date
                                response = requests.get(
                                    proj_url+proj["id"],
                                    headers= {'Authorization': token},
                                    verify=False
                                    )
                                
                                nodes = json.loads(response.content)['data']['nodes']
                                error_node = [n for n in nodes if n['status'] not in stat or n['export_status'] not in stat]
                        else:
                                status = "Not Synced"
                                error_note = f'Last Sync at {datetime_obj.strftime("%Y-%m-%d %H:%M:%S")}'
                else : 
                    response = requests.get(
                        proj_url+proj["id"],
                        headers= {'Authorization': token},
                        verify=False
                        )
                    nodes = json.loads(response.content)['data']['nodes']
                    error_node = [n for n in nodes if n['status'] not in stat or n['export_status'] not in stat]
            else :
                monitoring = requests.get(
                    proj_url+proj["id"]+"/monitoring?page_size=8&page=0",
                    headers= {'Authorization': token},
                    verify=False
                    )
                data = json.loads(monitoring.content)['data']
                for entry in data:
                    if entry.get('started_at'):  # Check if 'started_at' exists and is not None
                        date_sync = entry['started_at']
                        gmt7_datetime = convert_timezone_to_gmt7(date_sync)
                        datetime_obj = datetime.strptime(gmt7_datetime, "%a, %d %b %Y, %H:%M:%S %Z%z")
                         # Perform further operations with date_sync if needed
                        break  # Exit the loop once a valid 'started_at' is found
                status = "No Schedule"
                error_note = f'Last Sync at {datetime_obj.strftime("%Y-%m-%d %H:%M:%S")}'
        
        note, status = excel.write_line(proj, i+2, error_nodes=error_node, error_note=error_note, status=status)
        explored[proj["id"]] = (note, status)
        
    excel.save('summary_job.xlsx')

class ExcelWriter():
    def __init__(self):
        self.today = datetime.today().date().strftime("%d-%b-%y")
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.side = Side(border_style='thin', color='000000')
        header_border = Side(border_style='thin', color='000000')
        self.sheet["A1"] = self.today
        self.sheet.merge_cells('A1:L1')
        self.sheet["A1"].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
        self.sheet["A1"].border = Border(left=header_border, right=header_border, top=header_border, bottom=header_border)

        self.sheet.column_dimensions['B'].width = 11.43
        self.sheet.column_dimensions['C'].width = 11.43
        self.sheet.column_dimensions['F'].width = 54.57
        self.sheet.column_dimensions['G'].width = 56.71
        self.sheet.column_dimensions['H'].width = 12.14
        self.sheet.column_dimensions['I'].width = 16
        self.sheet.column_dimensions['J'].width = 74.14
        
        self.colors = {
            "Success": PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            "Failed": PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            "Not Synced": PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid'),
            "No Schedule": PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'),
            "": PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        }

    def write_line(self, proj, row_id,error_nodes=None, error_note=None,status=""):
        if error_note == None:
            error_note = ""
            if error_nodes is None:
                pass
            elif len(error_nodes) <= 5:
                for n in error_nodes:
                    if n["status"] != "SUCCESS":
                        error_note += f"{n['name']} --> {n['status']}\n"
                    else:
                        error_note += f"{n['name']} --> export {n['export_status']}\n"
            else:
                error_note = "error in more than 5 nodes"
                
            if error_nodes == None:
                pass
            elif len(error_nodes) > 0:
                status = "Failed"
            else:
                status = "Success"

        self.sheet["A" + str(row_id)].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        self.sheet["B" + str(row_id)] = self.today
        self.sheet["C" + str(row_id)] = datetime.today().strftime("%H:%M")
        self.sheet["F" + str(row_id)] = proj["init_name"]
        self.sheet["G" + str(row_id)] = proj["name"]
        self.sheet["H" + str(row_id)] = proj["loc"]
        self.sheet["I" + str(row_id)] = status
        self.sheet["I" + str(row_id)].fill = self.colors[status]
        self.sheet["J" + str(row_id)] = error_note
        self.sheet["L" + str(row_id)] = "Success" if status=="Success" else ""
        if status == "Success":
            self.sheet["L" + str(row_id)].fill = self.colors[status]
        
        for cell in self.sheet['A'+str(row_id):'L'+str(row_id)][0]:
            cell.border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        
        return error_note, status
        
    def save(self, name):
        self.wb.save(name)

if __name__ == '__main__':
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    load_dotenv(override=True)
    username = os.environ.get("USERNAME_DELMAN")
    password = os.environ.get("PASSWORD_DELMAN")
    token = get_token(username, password)
    excel = ExcelWriter()
    generate_excel(token, excel, *sys.argv[1:])